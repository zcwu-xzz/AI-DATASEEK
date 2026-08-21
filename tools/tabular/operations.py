#!/usr/bin/env python3
from __future__ import annotations

import argparse
import base64
import json
import math
import os
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd

MAX_CELLS = 2_000_000


def scalar(value: Any) -> Any:
    if value is None or isinstance(value, (str, bool, int)): return value
    if isinstance(value, (float, np.floating)): return float(value) if math.isfinite(float(value)) else None
    if isinstance(value, np.integer): return int(value)
    if isinstance(value, (pd.Timestamp, np.datetime64)): return str(value)
    if isinstance(value, dict): return {str(k): scalar(v) for k, v in value.items()}
    if isinstance(value, (list, tuple, np.ndarray)): return [scalar(v) for v in value]
    return str(value)


def response(operation: str, summary: dict[str, Any], *, artifacts: list[dict[str, Any]] | None = None, warnings: list[str] | None = None) -> dict[str, Any]:
    return scalar({"success":True,"answer_ready":True,"operation":operation,"summary":summary,"evidence":[],"artifacts":artifacts or [],"warnings":warnings or [],"provenance":{"tool":operation,"version":"1.0.0"},"recommended_next_tools":[]})


def output_path(raw: str) -> Path:
    root = Path(os.environ.get("AI_DATASEEK_OUTPUT_ROOT", "/home/ubuntu/output")).resolve()
    path = Path(raw).resolve()
    if root != path and root not in path.parents: raise ValueError(f"output path must be below {root}")
    path.parent.mkdir(parents=True, exist_ok=True)
    return path


def read_table(path: str, sheet: Any = None, header_row: int | None = 1, max_rows: int = 20000) -> pd.DataFrame:
    source, suffix = Path(path), Path(path).suffix.lower()
    header = None if header_row is None else header_row - 1
    if suffix in {".csv", ".tsv"}:
        frame = pd.read_csv(source, sep="\t" if suffix == ".tsv" else ",", header=header, nrows=max_rows)
    elif suffix in {".xlsx", ".xlsm", ".xls", ".ods"}:
        frame = pd.read_excel(source, sheet_name=0 if sheet is None else sheet, header=header, nrows=max_rows)
    else: raise ValueError("supported table formats are CSV, TSV, XLSX, XLSM, XLS and ODS")
    if not isinstance(frame, pd.DataFrame): raise ValueError("one explicit sheet is required")
    if frame.shape[0] * max(1, frame.shape[1]) > MAX_CELLS: raise ValueError("table exceeds the bounded cell limit")
    frame.columns = [str(column) for column in frame.columns]
    return frame


def artifact(path: Path) -> list[dict[str, Any]]:
    mime = "text/csv" if path.suffix.lower() == ".csv" else "application/json"
    return [{"path":str(path),"type":mime,"size_bytes":path.stat().st_size}]


def write_frame(frame: pd.DataFrame, raw: str) -> list[dict[str, Any]]:
    path = output_path(raw)
    if path.suffix.lower() == ".csv": frame.to_csv(path, index=False)
    elif path.suffix.lower() == ".json": frame.to_json(path, orient="records", force_ascii=False, date_format="iso")
    else: raise ValueError("table output must be .csv or .json")
    return artifact(path)


def workbook_inspect(a: dict[str, Any]) -> dict[str, Any]:
    path, suffix = Path(a["input_path"]), Path(a["input_path"]).suffix.lower()
    if suffix in {".csv", ".tsv"}:
        frame = read_table(str(path), max_rows=100)
        return response("workbook_inspect", {"format":suffix[1:],"sheets":[{"name":path.stem,"sample_rows":len(frame),"columns":list(frame.columns)}]})
    if suffix in {".xlsx", ".xlsm"}:
        import openpyxl
        book = openpyxl.load_workbook(path, read_only=False, data_only=False, keep_links=True)
        sheets=[]
        for sheet in book.worksheets:
            formulas=sum(1 for row in sheet.iter_rows() for cell in row if cell.data_type == "f")
            sheets.append({"name":sheet.title,"state":sheet.sheet_state,"max_row":sheet.max_row,"max_column":sheet.max_column,"merged_ranges":[str(r) for r in list(sheet.merged_cells.ranges)[:200]],"formula_count":formulas})
        names=list(book.defined_names)
        return response("workbook_inspect", {"format":suffix[1:],"sheets":sheets,"defined_names":names,"external_links":len(getattr(book,"_external_links", []))})
    excel = pd.ExcelFile(path)
    return response("workbook_inspect", {"format":suffix[1:],"sheets":[{"name":name} for name in excel.sheet_names]}, warnings=["detailed cell metadata is unavailable for this workbook format"])


def table_extract(a: dict[str, Any]) -> dict[str, Any]:
    frame=read_table(a["input_path"],a.get("sheet"),a.get("header_row",1),a.get("max_rows",5000))
    if a.get("cell_range"):
        from openpyxl.utils.cell import range_boundaries
        left,top,right,bottom=range_boundaries(a["cell_range"]); frame=frame.iloc[max(0,top-2):bottom-1,left-1:right]
    artifacts=write_frame(frame,a["output_path"]) if a.get("output_path") else []
    return response("table_extract", {"rows":len(frame),"columns":list(frame.columns),"preview":frame.head(20).to_dict("records")}, artifacts=artifacts)


def column_profile(series: pd.Series) -> dict[str, Any]:
    values=series.dropna(); item={"dtype":str(series.dtype),"missing":int(series.isna().sum()),"unique":int(values.nunique()),"examples":scalar(values.head(5).tolist())}
    numeric=pd.to_numeric(values,errors="coerce").dropna()
    if len(numeric) >= max(1,len(values)//2): item["numeric"]={"min":numeric.min(),"max":numeric.max(),"mean":numeric.mean(),"median":numeric.median()}
    dates=pd.to_datetime(values,errors="coerce").dropna()
    if len(dates) >= max(1,len(values)//2): item["time_range"]=[dates.min(),dates.max()]
    return scalar(item)


def table_profile(a: dict[str, Any]) -> dict[str, Any]:
    frame=read_table(a["input_path"],a.get("sheet"),a.get("header_row",1),a.get("max_rows",20000))
    return response("table_profile", {"rows":len(frame),"columns":{c:column_profile(frame[c]) for c in frame},"duplicate_rows":int(frame.duplicated().sum())})


def schema_infer(a: dict[str, Any]) -> dict[str, Any]:
    raw=read_table(a["input_path"],a.get("sheet"),None,a.get("sample_rows",50)); scores=[]
    for index,row in raw.head(10).iterrows():
        text=row.notna().sum(); unique=row.dropna().astype(str).nunique(); scores.append({"row":int(index)+1,"score":float(text+unique/max(1,text))})
    candidate=max(scores,key=lambda x:x["score"])["row"] if scores else None
    frame=read_table(a["input_path"],a.get("sheet"),candidate,a.get("sample_rows",50)) if candidate else raw
    roles={}
    for c in frame:
        p=column_profile(frame[c]); roles[c]="numeric" if "numeric" in p else "datetime" if "time_range" in p else "identifier" if p["unique"]==len(frame) else "categorical"
    return response("table_schema_infer", {"header_row_candidate":candidate,"header_scores":scores,"column_roles":roles}, warnings=["header inference is heuristic; pass header_row explicitly before authoritative analysis"])


def apply_filters(frame: pd.DataFrame, filters: list[dict[str, Any]]) -> pd.DataFrame:
    for rule in filters:
        c,op,v=rule["column"],rule["operator"],rule["value"]
        if c not in frame: raise ValueError(f"unknown filter column: {c}")
        s=frame[c]
        masks={"eq":lambda:s==v,"ne":lambda:s!=v,"gt":lambda:s>v,"ge":lambda:s>=v,"lt":lambda:s<v,"le":lambda:s<=v,"contains":lambda:s.astype(str).str.contains(str(v),case=False,na=False,regex=False),"in":lambda:s.isin(v if isinstance(v,list) else [v])}
        frame=frame[masks[op]()].copy()
    return frame


def filter_aggregate(a: dict[str, Any]) -> dict[str, Any]:
    frame=apply_filters(read_table(a["input_path"],a.get("sheet")),a.get("filters",[])); groups=a.get("group_by",[]); aggs=a["aggregations"]
    for c in [*groups,*aggs]:
        if c not in frame: raise ValueError(f"unknown column: {c}")
    out=frame.groupby(groups,dropna=False).agg(aggs).reset_index() if groups else pd.DataFrame([{f"{c}_{m}":getattr(frame[c],m)() for c,m in aggs.items()}])
    return response("table_filter_aggregate", {"input_rows_after_filter":len(frame),"result_rows":len(out),"preview":out.head(20).to_dict("records")}, artifacts=write_frame(out,a["output_path"]))


def join_compare(a: dict[str, Any]) -> dict[str, Any]:
    left,right=read_table(a["left_path"],a.get("left_sheet")),read_table(a["right_path"],a.get("right_sheet")); keys=a["keys"]
    for key in keys:
        if key not in left or key not in right: raise ValueError(f"join key is missing: {key}")
    merged=left.merge(right,on=keys,how=a.get("how","outer"),suffixes=("_left","_right"),indicator=True,validate="many_to_many")
    changed=np.zeros(len(merged),dtype=bool)
    matched=merged["_merge"].eq("both").to_numpy()
    for c in sorted((set(left.columns)&set(right.columns))-set(keys)):
        differs=~((merged[f"{c}_left"]==merged[f"{c}_right"]) | (merged[f"{c}_left"].isna()&merged[f"{c}_right"].isna()))
        changed |= matched & differs.to_numpy()
    merged["_changed"]=changed
    counts=merged["_merge"].value_counts().to_dict(); counts["changed"]=int(changed.sum())
    return response("table_join_compare", {"counts":counts,"rows":len(merged)}, artifacts=write_frame(merged,a["output_path"]))


def formula_audit(a: dict[str, Any]) -> dict[str, Any]:
    import openpyxl
    path=Path(a["input_path"])
    if path.suffix.lower() not in {".xlsx",".xlsm"}: raise ValueError("formula audit supports XLSX and XLSM")
    book=openpyxl.load_workbook(path,data_only=False,read_only=False,keep_links=True); formulas=[]; total=0
    for sheet in book.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.data_type=="f":
                    total+=1
                    if len(formulas)<a.get("max_formulas",2000): formulas.append({"sheet":sheet.title,"cell":cell.coordinate,"formula":cell.value})
    cached=openpyxl.load_workbook(path,data_only=True,read_only=False)
    missing=sum(1 for item in formulas if cached[item["sheet"]][item["cell"]].value is None)
    return response("workbook_formula_audit", {"formula_count":total,"formulas":formulas,"sampled_formula_count":len(formulas),"sampled_missing_cached_values":missing,"external_links":len(getattr(book,"_external_links",[]))}, warnings=["formulas are audited but not recalculated; cached values may be stale"])


def visualize(a: dict[str, Any]) -> dict[str, Any]:
    import matplotlib; matplotlib.use("Agg"); import matplotlib.pyplot as plt
    frame=read_table(a["input_path"],a.get("sheet"),max_rows=a.get("max_rows",20000)); chart,x,ys=a["chart"],a.get("x"),a.get("y",[])
    for c in ([x] if x else [])+ys+([a["group"]] if a.get("group") else []):
        if c not in frame: raise ValueError(f"unknown chart column: {c}")
    fig,ax=plt.subplots(figsize=(10,6),constrained_layout=True)
    if chart=="correlation_heatmap":
        values=frame[ys].apply(pd.to_numeric,errors="coerce") if ys else frame.select_dtypes(include="number"); image=ax.imshow(values.corr(),vmin=-1,vmax=1,cmap="coolwarm"); ax.set_xticks(range(len(values.columns)),values.columns,rotation=45,ha="right"); ax.set_yticks(range(len(values.columns)),values.columns); fig.colorbar(image,ax=ax)
    elif chart=="histogram": frame[ys[0]].plot.hist(ax=ax,bins=30)
    elif chart=="box": frame[ys].plot.box(ax=ax)
    elif chart=="scatter": ax.scatter(frame[x],frame[ys[0]],s=18,alpha=.7)
    else: frame.plot(x=x,y=ys,kind=chart,ax=ax)
    ax.set_title(a.get("title") or ""); ax.grid(chart in {"line","scatter"},alpha=.25)
    path=output_path(a["output_path"]); fig.savefig(path,dpi=160); plt.close(fig)
    return response("tabular_visualize", {"chart":chart,"rows_used":len(frame),"x":x,"y":ys}, artifacts=[{"path":str(path),"type":"image/png","size_bytes":path.stat().st_size}])


def chart_recommend(a: dict[str, Any]) -> dict[str, Any]:
    frame=read_table(a["input_path"],a.get("sheet"),max_rows=500); profiles={c:column_profile(frame[c]) for c in frame}; numeric=[c for c,p in profiles.items() if "numeric" in p]; dates=[c for c,p in profiles.items() if "time_range" in p]; categorical=[c for c in frame if c not in numeric and c not in dates]
    recommendations=[]
    if dates and numeric: recommendations.append({"chart":"line","x":dates[0],"y":numeric[:4],"reason":"time and numeric columns"})
    if categorical and numeric: recommendations.append({"chart":"bar","x":categorical[0],"y":numeric[:2],"reason":"category and numeric columns"})
    if len(numeric)>=2: recommendations += [{"chart":"scatter","x":numeric[0],"y":[numeric[1]],"reason":"two numeric columns"},{"chart":"correlation_heatmap","y":numeric[:12],"reason":"multiple numeric columns"}]
    if numeric: recommendations.append({"chart":"histogram","y":[numeric[0]],"reason":"numeric distribution"})
    return response("table_chart_recommend", {"recommendations":recommendations,"numeric":numeric,"datetime":dates,"categorical":categorical})


def workbook_quality_audit(a: dict[str, Any]) -> dict[str, Any]:
    """Audit bounded worksheet quality without mutating or recalculating a workbook."""
    import openpyxl

    path = Path(a["input_path"])
    suffix = path.suffix.lower()
    if suffix not in {".xlsx", ".xlsm"}:
        # Legacy XLS is read through pandas, which cannot expose cell-level
        # metadata consistently across xlrd versions.
        excel = pd.ExcelFile(path)
        sheets = []
        for name in excel.sheet_names[: a.get("max_sheets", 30)]:
            frame = read_table(str(path), name, max_rows=a.get("max_rows", 20000))
            sheets.append({"name": name, "rows": len(frame), "columns": len(frame.columns), "missing_cells": int(frame.isna().sum().sum()), "duplicate_rows": int(frame.duplicated().sum()), "warnings": ["cell-level formatting and formula errors are unavailable for XLS"]})
        return response("workbook_quality_audit", {"format": suffix[1:], "sheets": sheets, "sheet_count": len(excel.sheet_names)}, warnings=["XLS quality audit is table-level because legacy cell metadata is unavailable"])

    book = openpyxl.load_workbook(path, read_only=False, data_only=False, keep_links=True)
    max_sheets = a.get("max_sheets", 30)
    max_cells = a.get("max_cells", 200000)
    sheets = []
    scanned = 0
    for sheet in book.worksheets[:max_sheets]:
        nonempty = 0
        formulas = 0
        errors = 0
        blank_rows = 0
        header_values: list[Any] = []
        for row_index, row in enumerate(sheet.iter_rows(), start=1):
            if scanned >= max_cells:
                break
            values = [cell.value for cell in row]
            scanned += len(values)
            nonempty += sum(value is not None for value in values)
            if not any(value is not None for value in values):
                blank_rows += 1
            if row_index == 1:
                header_values = values
            for cell in row:
                if cell.data_type == "f":
                    formulas += 1
                if cell.data_type == "e" or (isinstance(cell.value, str) and cell.value.startswith("#")):
                    errors += 1
        headers = [str(value).strip() for value in header_values if value is not None and str(value).strip()]
        duplicates = sorted({value for value in headers if headers.count(value) > 1})
        total_cells = max(1, min(sheet.max_row * max(1, sheet.max_column), max_cells))
        sheets.append({"name": sheet.title, "state": sheet.sheet_state, "rows": sheet.max_row, "columns": sheet.max_column, "nonempty_cells": nonempty, "missing_ratio": round(max(0, total_cells - nonempty) / total_cells, 6), "blank_rows": blank_rows, "formula_count": formulas, "error_cell_count": errors, "duplicate_headers": duplicates, "merged_range_count": len(sheet.merged_cells.ranges)})
    warnings = []
    if len(book.worksheets) > max_sheets:
        warnings.append(f"only the first {max_sheets} worksheets were audited")
    if scanned >= max_cells:
        warnings.append(f"cell scan capped at {max_cells} cells")
    return response("workbook_quality_audit", {"format": suffix[1:], "sheet_count": len(book.worksheets), "sheets": sheets, "scanned_cells": scanned}, warnings=warnings)


def workbook_concat_sheets(a: dict[str, Any]) -> dict[str, Any]:
    """Append explicitly selected worksheets and retain their source sheet."""
    path = Path(a["input_path"])
    excel = pd.ExcelFile(path)
    requested = a.get("sheets") or excel.sheet_names
    unknown = [name for name in requested if name not in excel.sheet_names]
    if unknown:
        raise ValueError(f"unknown worksheet: {unknown[0]}")
    frames = []
    for name in requested[: a.get("max_sheets", 20)]:
        frame = read_table(str(path), name, a.get("header_row", 1), a.get("max_rows_per_sheet", 20000)).copy()
        frame.insert(0, "_source_sheet", name)
        frames.append(frame)
    if not frames:
        raise ValueError("at least one worksheet is required")
    combined = pd.concat(frames, ignore_index=True, sort=False)
    artifacts = write_frame(combined, a["output_path"])
    return response("workbook_concat_sheets", {"sheets": requested[: a.get("max_sheets", 20)], "rows": len(combined), "columns": list(combined.columns), "preview": combined.head(20).to_dict("records")}, artifacts=artifacts)


def table_pivot(a: dict[str, Any]) -> dict[str, Any]:
    frame = read_table(a["input_path"], a.get("sheet"), a.get("header_row", 1), a.get("max_rows", 50000))
    index = a["index"] if isinstance(a["index"], list) else [a["index"]]
    columns = a.get("columns")
    values = a.get("values")
    for name in [*index, *(columns if isinstance(columns, list) else [columns] if columns else []), *(values if isinstance(values, list) else [values] if values else [])]:
        if name not in frame:
            raise ValueError(f"unknown pivot column: {name}")
    pivot = pd.pivot_table(frame, index=index, columns=columns, values=values, aggfunc=a.get("aggfunc", "mean"), fill_value=a.get("fill_value"), dropna=False).reset_index()
    pivot.columns = ["_".join(str(part) for part in column if str(part) != "") if isinstance(column, tuple) else str(column) for column in pivot.columns]
    artifacts = write_frame(pivot, a["output_path"])
    return response("table_pivot", {"rows": len(pivot), "columns": list(pivot.columns), "index": index, "values": values, "aggfunc": a.get("aggfunc", "mean"), "preview": pivot.head(20).to_dict("records")}, artifacts=artifacts)


FUNCTIONS={"workbook_inspect":workbook_inspect,"table_extract":table_extract,"table_profile":table_profile,"table_schema_infer":schema_infer,"table_filter_aggregate":filter_aggregate,"table_join_compare":join_compare,"workbook_formula_audit":formula_audit,"workbook_quality_audit":workbook_quality_audit,"workbook_concat_sheets":workbook_concat_sheets,"table_pivot":table_pivot,"tabular_visualize":visualize,"table_chart_recommend":chart_recommend}


def main() -> int:
    parser=argparse.ArgumentParser(); parser.add_argument("tool"); parser.add_argument("payload"); args=parser.parse_args()
    try: output=FUNCTIONS[args.tool](json.loads(base64.urlsafe_b64decode(args.payload+"="*(-len(args.payload)%4))))
    except Exception as exc: output={"success":False,"answer_ready":True,"operation":args.tool,"error":f"{type(exc).__name__}: {exc}","warnings":[]}
    print(json.dumps(scalar(output),ensure_ascii=False,allow_nan=False)); return 0 if output["success"] else 1


if __name__=="__main__": raise SystemExit(main())
